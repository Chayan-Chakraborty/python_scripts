import os
from contextlib import contextmanager

try:
    from sshtunnel import SSHTunnelForwarder
except ImportError:
    SSHTunnelForwarder = None

try:
    import mysql.connector
except ImportError:
    mysql = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# ================= ENV CONFIG ================= #
TARGET_ENV = os.getenv("TARGET_ENV", "local").lower()

ENV_CONFIG = {
    "local": {
        "MYSQL_USER": "root",
        "MYSQL_PASSWORD": "Chayan@1340",
        "MYSQL_HOST": "localhost",
        "MYSQL_PORT": 3306,
        "MYSQL_DATABASE": "data_warehouse",
    }
}
# ================= ENV CONFIG END ================= #

if TARGET_ENV not in ENV_CONFIG:
    raise ValueError("❌ TARGET_ENV must be one of: local, dev, stage, prod")

_cfg = ENV_CONFIG[TARGET_ENV]
MYSQL_HOST = _cfg["MYSQL_HOST"]
MYSQL_PORT = _cfg["MYSQL_PORT"]
MYSQL_USER = _cfg["MYSQL_USER"]
MYSQL_PASSWORD = _cfg["MYSQL_PASSWORD"]
MYSQL_DATABASE = _cfg["MYSQL_DATABASE"]
SSH_HOST = _cfg.get("SSH_HOST")
SSH_PORT = _cfg.get("SSH_PORT", 22)
SSH_USER = _cfg.get("SSH_USER")
SSH_PASSWORD = _cfg.get("SSH_PASSWORD")
SSH_PRIVATE_KEY = _cfg.get("SSH_KEY_PATH") or _cfg.get("SSH_PRIVATE_KEY")
SSH_PRIVATE_KEY_PASSWORD = _cfg.get("SSH_KEY_PASSPHRASE") or _cfg.get("SSH_PRIVATE_KEY_PASSWORD")


@contextmanager
def ssh_tunnel_if_needed(mysql_cfg: dict):
    use_ssh = bool(SSH_HOST and SSH_USER)

    if not use_ssh:
        yield mysql_cfg
        return

    if SSHTunnelForwarder is None:
        raise RuntimeError(
            "sshtunnel is required for SSH-based MySQL connections. "
            "Install with: pip install sshtunnel"
        )

    ssh_kwargs = {
        "ssh_username": SSH_USER,
    }
    if SSH_PASSWORD:
        ssh_kwargs["ssh_password"] = SSH_PASSWORD
    if SSH_PRIVATE_KEY:
        ssh_kwargs["ssh_pkey"] = SSH_PRIVATE_KEY
    if SSH_PRIVATE_KEY_PASSWORD:
        ssh_kwargs["ssh_private_key_password"] = SSH_PRIVATE_KEY_PASSWORD

    server = SSHTunnelForwarder(
        (SSH_HOST, SSH_PORT),
        remote_bind_address=(mysql_cfg["host"], mysql_cfg["port"]),
        **ssh_kwargs,
    )
    server.start()
    try:
        forwarded_cfg = dict(mysql_cfg)
        forwarded_cfg["host"] = "127.0.0.1"
        forwarded_cfg["port"] = server.local_bind_port
        yield forwarded_cfg
    finally:
        server.stop()


@contextmanager
def mysql_connection():
    if not hasattr(mysql, "connector"):
        raise RuntimeError(
            "mysql-connector-python is required for MySQL connections. "
            "Install with: pip install mysql-connector-python"
        )

    base_cfg = {
        "host": MYSQL_HOST,
        "port": MYSQL_PORT,
        "user": MYSQL_USER,
        "password": MYSQL_PASSWORD,
        "database": MYSQL_DATABASE,
    }

    with ssh_tunnel_if_needed(base_cfg) as effective_cfg:
        conn = mysql.connector.connect(
            host=effective_cfg["host"],
            port=effective_cfg["port"],
            user=effective_cfg["user"],
            password=effective_cfg["password"],
            database=effective_cfg["database"],
            charset="utf8mb4",
        )
        try:
            yield conn
        finally:
            conn.close()


def fetch_event_names():
    with mysql_connection() as conn:
        with conn.cursor(dictionary=True) as cur:
            cur.execute("SELECT event_name FROM events")
            rows = cur.fetchall()
            return [row["event_name"] for row in rows if row.get("event_name") is not None]


def fetch_event_names_from_excel(
    excel_path: str = "/Users/chayanchakraborty/Documents/python_new_dir/git-operations/Event_Name_List.xlsx",
):
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to read Excel files. "
            "Install with: pip install openpyxl"
        )

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        event_names = []
        seen = set()

        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue

            headers = [str(col).strip().lower().replace("_", " ") if col is not None else "" for col in header_row]
            if "event name" not in headers:
                continue

            event_idx = headers.index("event name")
            for row in rows:
                if row is None or event_idx >= len(row):
                    continue
                value = row[event_idx]
                if value is None:
                    continue

                event_name = str(value).strip()
                if not event_name:
                    continue

                key = event_name.lower()
                if key in seen:
                    continue
                seen.add(key)
                event_names.append(event_name)

        return event_names
    finally:
        workbook.close()


def get_missing_events_in_db(excel_event_names, db_event_names):
    db_keys = {str(name).strip().lower() for name in db_event_names if name is not None}
    missing = []
    seen = set()

    for name in excel_event_names:
        if name is None:
            continue
        event_name = str(name).strip()
        if not event_name:
            continue
        key = event_name.lower()
        if key in db_keys or key in seen:
            continue
        seen.add(key)
        missing.append(event_name)

    return missing


def insert_missing_events(missing_event_names):
    if not missing_event_names:
        return 0

    with mysql_connection() as conn:
        with conn.cursor() as cur:
            values = [(name, 1, None) for name in missing_event_names]
            cur.executemany(
                "INSERT INTO events (event_name, is_active, start_date, created_at, updated_at) "
                "VALUES (%s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)",
                values,
            )
        conn.commit()

    return len(missing_event_names)


if __name__ == "__main__":
    db_event_names = fetch_event_names()
    excel_event_names = fetch_event_names_from_excel()
    missing_event_names = get_missing_events_in_db(excel_event_names, db_event_names)

    print(f"DB events: {len(db_event_names)}")
    print(f"Excel events: {len(excel_event_names)}")
    print(f"Missing in DB: {len(missing_event_names)}")

    inserted_count = insert_missing_events(missing_event_names)
    print(f"Inserted missing events: {inserted_count}")

