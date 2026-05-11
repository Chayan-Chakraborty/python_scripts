from openpyxl import Workbook, load_workbook

INPUT_EXCEL_PATH = "/Users/chayanchakraborty/Documents/python_new_dir/git-operations/CleverTap_Event_Hierarchy.xlsx"
OUTPUT_EXCEL_PATH = "/Users/chayanchakraborty/Documents/python_new_dir/git-operations/Event_Name_List.xlsx"


def fetch_event_names_from_excel(excel_path: str):
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        event_names = []
        seen = set()

        table_header_keywords = {
            "event name",
            "key properties",
            "trigger",
            "description",
            "sdk call",
            "location",
            "purpose",
            "layer",
            "class / file",
            "method",
            "target",
        }

        for sheet in workbook.worksheets:
            event_idx = None
            for row in sheet.iter_rows(values_only=True):
                if row is None:
                    continue

                cells = [str(col).strip() if col is not None else "" for col in row]
                normalized_cells = [cell.lower().replace("_", " ") for cell in cells]
                non_empty_cells = [cell for cell in normalized_cells if cell]

                if "event name" in normalized_cells:
                    event_idx = normalized_cells.index("event name")
                    continue

                if event_idx is None:
                    continue

                if (
                    len(non_empty_cells) >= 2
                    and any(cell in table_header_keywords for cell in non_empty_cells)
                    and "event name" not in non_empty_cells
                ):
                    event_idx = None
                    continue

                if event_idx >= len(cells):
                    continue

                event_name = cells[event_idx].strip()
                if not event_name or event_name.lower() == "event name":
                    continue

                has_context_cols = any(
                    cells[idx].strip() for idx in range(len(cells)) if idx != event_idx
                )
                if not has_context_cols:
                    continue

                key = event_name.lower()
                if key in seen:
                    continue
                seen.add(key)
                event_names.append(event_name)

        return event_names
    finally:
        workbook.close()


def write_event_names_excel(event_names, output_path: str):
    out_wb = Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "Event Names"
    out_sheet.append(["Event Name", "is_active", "start_date"])
    for event_name in event_names:
        out_sheet.append([event_name, 1, "2026-01-01"])
    out_wb.save(output_path)


def main():
    event_names = fetch_event_names_from_excel(INPUT_EXCEL_PATH)
    write_event_names_excel(event_names, OUTPUT_EXCEL_PATH)
    print(f"Total events extracted: {len(event_names)}")
    print(f"Output file created: {OUTPUT_EXCEL_PATH}")


if __name__ == "__main__":
    main()
