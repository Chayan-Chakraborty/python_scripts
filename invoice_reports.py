import json
import mysql.connector
import pandas as pd
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
import logging
# Database configurations
db_configs = {
    'local':{
        'user': 'root',
        'password': 'secretpassword',
        'host': 'localhost',
        'database': 'garuda',
        'port' : 3306
    },
    'prod':{
        'user': 'cron_read_user',
        'password': 'J5drbt7yHT8NAj4',
        'host': 'solstice-prod-rds.cjkcxc4q21f9.ap-south-1.rds.amazonaws.com',
        'database': 'garuda',
        'port' : 3306
    }
}

current_date = datetime.today().strftime('%Y-%m-%d')

def get_db_config(name):
    return db_configs.get(name)
# Email configuration
email_configs = {
    'local': {
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,  # For TLS
        'username': 'app_store_dev@pidilite.vc',
        'password': 'gpbs zfep pdwk awgv',
        'from_email': 'app_store_dev@pidilite.vc',
        'to_emails': ['saransh.agrawal@pidilite.vc'],  # List of recipients
        'subject': f'Invoice Data : {current_date}',
        'body': 'Please find the attached Excel file with the data.'
    },
    'prod': {
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,  # For TLS
        'username': 'app_store_dev@pidilite.vc',
        'password': 'gpbs zfep pdwk awgv',
        'from_email': 'app_store_dev@pidilite.vc',
        #'to_emails': ['ahona@pidilite.vc', 'anil@pidilite.vc','b.mustafa@pidilite.vc','sumit.singh@pidilite.vc','chinmay.jaripatke@pidilite.vc','deepak.agarwal@pidilite.vc','anurag@pidilite.vc'],  # List of recipients
        'to_emails': ['chayan.chakraborty@pidilite.vc'],
        'subject': f'Invoice Data : {current_date}',
        'body': 'Please find the attached Excel file with the data.'
    }
}
def get_emial_config(name):
    return email_configs.get(name)
# Function to fetch data from a database
def fetch_data_from_db(db_config, query):
    try:
        logging.info(f"Connecting to database: {db_config['database']} at {db_config['host']}:{db_config['port']}")
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)
        cursor.execute(query)
        data = cursor.fetchall()
        cursor.close()
        connection.close()
        return data
    except mysql.connector.Error as err:
        logging.error(f"Error: {err}")
        return []
# Function to generate an Excel file from data
def generate_excel(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Column headers
    headers = ['Invoice Number', 'Order Id', 'Shipment Id', 'Invoice Date', 'Contractor Name', 'Contractor Phone Number',
               'Invoice Amount']
    # Write headers to the first row
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)
    # Writing data to Excel cell by cell
    for row_index, row_data in enumerate(data, start=2):
        # print('Row data', row_data)
        items = json.loads(row_data['additional_info'])['items']
        total_amount = sum(item["totalAmount"] for item in items)
        # print(total_amount)
        col_index = 1
        for cell_key, cell_value in row_data.items():
            if cell_key != 'additional_info':
                # print('Writing', col_index, cell_key, cell_value)
                ws.cell(row=row_index, column=col_index, value=cell_value)
                col_index += 1
        # Write total amount to the last column
        ws.cell(row=row_index, column=len(headers), value=total_amount)
    # Save the workbook
    wb.save(filename)
# Function to send an email with the Excel attachment
def send_email_with_attachment(filename, environment_name):
    email_config = get_emial_config(environment_name)
    logging.info(f"Connecting to database: {email_config['username']} at {email_config['from_email']}:{email_config['to_emails']}")
    msg = MIMEMultipart()
    msg['From'] = email_config['from_email']
    msg['To'] = ', '.join(email_config['to_emails'])
    msg['Subject'] = email_config['subject']
    msg.attach(MIMEText(email_config['body'], 'plain'))
    attachment = open(filename, 'rb')
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={filename}')
    msg.attach(part)
    attachment.close()
    server = smtplib.SMTP(email_config['smtp_server'], email_config['smtp_port'])
    server.starttls()
    server.login(email_config['username'], email_config['password'])
    text = msg.as_string()
    server.sendmail(email_config['from_email'], email_config['to_emails'], text)
    server.quit()
def main(environment_name):
    #query = "SELECT i.external_id, s.order_external_id AS order_external_id, s.external_id AS shipment_external_id, i.invoice_date, i.additional_info, u.name, u.username FROM garuda.shipment s INNER JOIN garuda.invoice i ON i.entity_external_id = s.external_id INNER JOIN shield.app_user_v2 u ON s.user_id = u.id WHERE i.invoice_date >= DATE_FORMAT(CURDATE(), '%Y-%m-01') AND i.invoice_date <= CURDATE() order by i.external_id desc;"
    query = "SELECT i.external_id, s.order_external_id AS order_external_id, s.external_id AS shipment_external_id, i.invoice_date, i.additional_info, u.name, u.username FROM garuda.shipment s INNER JOIN garuda.invoice i ON i.entity_external_id = s.external_id INNER JOIN shield.app_user_v2 u ON s.user_id = u.id WHERE i.invoice_date >= DATE_FORMAT(CURDATE(), '%Y-08-01') AND i.invoice_date <= DATE_FORMAT(CURDATE(), '%Y-09-01') order by i.external_id desc;"
    all_data = []
    db_config = get_db_config(environment_name)
    # Fetch data from each database
    data = fetch_data_from_db(db_config, query)
    # print(data)
    all_data.extend(data)
    filename = 'Invoice-Data.xlsx'
    generate_excel(all_data, filename)
    send_email_with_attachment(filename, environment_name)
    print('Email sent successfully with the Excel attachment.')
# Main script
# if __name__ == "__main__":
#     import argparse
#     parser = argparse.ArgumentParser(description='Process environment name.')
#     parser.add_argument('environment_name', type=str, help='Environment name (e.g., prod, uat, dev)')
#     args = parser.parse_args()
#     main(args.environment_name)
main('prod')
